"""
Stock Validation App
=====================
Compares stock across three sources — Marketplace (e.g. Shopee), Platform, and
Warehouse — and flags discrepancies.

Business rules implemented:
  1. SKU universe = Marketplace file (SKU column F; falls back to Parent SKU
     column E when F is blank).
  2. Warehouse stock is the anchor / source of truth. When Marketplace or
     Platform disagree with Warehouse, Warehouse's number is treated as correct.
  3. Every SKU is checked against both Warehouse and Platform; mismatches are
     flagged with a status indicator and the size of the discrepancy.
  4. All SKUs with at least one discrepancy are compiled into a single
     downloadable "Result" file. A separate "Working Process" file contains
     the full line-by-line comparison for every SKU (audit trail).

Run locally with:
    streamlit run stock_validation_app.py
"""

import io
import re
import numpy as np
import pandas as pd
import streamlit as st

# ============================================================================
# PAGE CONFIG
# ============================================================================
st.set_page_config(
    page_title="Stock Validation Automation",
    page_icon="📦",
    layout="wide",
)

st.markdown(
    """
    <style>
    .block-container {padding-top: 2rem;}
    div[data-testid="stMetric"] {
        background-color: #F7F9FC;
        border: 1px solid #E5E9F0;
        border-radius: 10px;
        padding: 14px 16px;
    }
    div[data-testid="stMetric"] label,
    div[data-testid="stMetric"] label p,
    div[data-testid="stMetricLabel"],
    div[data-testid="stMetricLabel"] p {
        color: #4A5568 !important;
    }
    div[data-testid="stMetric"] div[data-testid="stMetricValue"] {
        color: #12181F !important;
    }
    div[data-testid="stMetric"] div[data-testid="stMetricDelta"] {
        color: #C13A3A !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


# ============================================================================
# CORE LOGIC
# ============================================================================
def parse_marketplace_file_shopee(file) -> pd.DataFrame:
    """Parse a Shopee-style marketplace export. SKU = column F, falls back to
    Parent SKU column E when F is blank. Auto-detects header/data rows."""
    raw = pd.read_excel(file, header=None, dtype=object)

    header_row = None
    for i in range(min(20, len(raw))):
        row_vals = raw.iloc[i].astype(str).str.strip().str.lower()
        if (row_vals == "parent sku").any() and (row_vals == "sku").any():
            header_row = i
            break
    if header_row is None:
        raise ValueError(
            "Could not find the header row in the Marketplace file "
            "(expected columns labelled 'Parent SKU' and 'SKU')."
        )

    header_vals = raw.iloc[header_row].astype(str).str.strip().str.lower()

    def find_col(label):
        matches = header_vals[header_vals == label].index
        if len(matches) == 0:
            raise ValueError(f"Could not find required column '{label}' in Marketplace file.")
        return matches[0]

    col_product_name = find_col("product name")
    col_parent_sku = find_col("parent sku")
    col_sku = find_col("sku")
    col_stock = find_col("stock")
    col_product_id = find_col("product id")

    data = raw.iloc[header_row + 1:].copy()
    pid = pd.to_numeric(data[col_product_id], errors="coerce")
    data = data[pid.notna()]

    out = pd.DataFrame({
        "Product Name": data[col_product_name],
        "Parent SKU": data[col_parent_sku].astype(str).str.strip(),
        "SKU (F col)": data[col_sku].astype(str).str.strip(),
        "Marketplace_Stock": pd.to_numeric(data[col_stock], errors="coerce").fillna(0),
    })
    out["Parent SKU"] = out["Parent SKU"].replace({"nan": np.nan, "None": np.nan, "": np.nan})
    out["SKU (F col)"] = out["SKU (F col)"].replace({"nan": np.nan, "None": np.nan, "": np.nan})

    out["SKU"] = out["SKU (F col)"].where(out["SKU (F col)"].notna(), out["Parent SKU"])
    out["SKU_Source"] = np.where(out["SKU (F col)"].notna(), "SKU (F)", "Parent SKU (E)")
    out = out.dropna(subset=["SKU"]).reset_index(drop=True)

    out = out.groupby("SKU", as_index=False).agg({
        "Product Name": "first",
        "SKU_Source": "first",
        "Marketplace_Stock": "sum",
    })
    return out[["SKU", "Product Name", "SKU_Source", "Marketplace_Stock"]]


def parse_marketplace_file_tiktok(file) -> pd.DataFrame:
    """Parse a TikTok Shop-style marketplace export (Indonesian template).
    SKU = column H ('SKU Penjual'). Stock = column G ('Kuantitas').
    Prefers the Indonesian header row; falls back to the English machine-key
    row ('seller_sku' / 'quantity') if no Indonesian header is found.
    Auto-detects header/data rows (data rows = where Product ID parses as a number)."""
    raw = pd.read_excel(file, header=None, dtype=object)

    def find_header_row(id_label, qty_label):
        for i in range(min(20, len(raw))):
            row_vals = raw.iloc[i].astype(str).str.strip().str.lower()
            if (row_vals == id_label).any() and (row_vals == qty_label).any():
                return i
        return None

    header_row = find_header_row("sku penjual", "kuantitas")
    if header_row is None:
        header_row = find_header_row("seller_sku", "quantity")
    if header_row is None:
        raise ValueError(
            "Could not find the header row in the TikTok Marketplace file "
            "(expected 'SKU Penjual'/'Kuantitas' or 'seller_sku'/'quantity' labels)."
        )

    header_vals = raw.iloc[header_row].astype(str).str.strip().str.lower()

    def find_col(*labels):
        for label in labels:
            matches = header_vals[header_vals == label].index
            if len(matches):
                return matches[0]
        raise ValueError(f"Could not find any of {labels} in TikTok Marketplace file.")

    col_product_name = find_col("product name", "nama produk", "product_name")
    col_sku = find_col("sku penjual", "seller_sku")
    col_stock = find_col("kuantitas", "quantity")
    col_product_id = find_col("id produk", "product_id")

    data = raw.iloc[header_row + 1:].copy()
    pid = pd.to_numeric(data[col_product_id], errors="coerce")
    data = data[pid.notna()]

    out = pd.DataFrame({
        "Product Name": data[col_product_name],
        "SKU": data[col_sku].astype(str).str.strip(),
        "Marketplace_Stock": pd.to_numeric(data[col_stock], errors="coerce").fillna(0),
    })
    out["SKU"] = out["SKU"].replace({"nan": np.nan, "None": np.nan, "": np.nan})
    out = out.dropna(subset=["SKU"]).reset_index(drop=True)
    out["SKU_Source"] = "SKU Penjual (H)"

    out = out.groupby("SKU", as_index=False).agg({
        "Product Name": "first",
        "SKU_Source": "first",
        "Marketplace_Stock": "sum",
    })
    return out[["SKU", "Product Name", "SKU_Source", "Marketplace_Stock"]]


def parse_marketplace_file_lazada(file) -> pd.DataFrame:
    """Parse a Lazada-style marketplace export (Indonesian template).
    SKU = column M ('Seller SKU'). Stock = column H ('Jumlah Stok').
    Auto-detects header/data rows (data rows = where Product ID parses as a number)."""
    raw = pd.read_excel(file, header=None, dtype=object)

    header_row = None
    for i in range(min(20, len(raw))):
        row_vals = raw.iloc[i].astype(str).str.strip().str.lower()
        if (row_vals == "seller sku").any() and (row_vals == "jumlah stok").any():
            header_row = i
            break
    if header_row is None:
        raise ValueError(
            "Could not find the header row in the Lazada Marketplace file "
            "(expected 'Seller SKU' / 'Jumlah Stok' labels)."
        )

    header_vals = raw.iloc[header_row].astype(str).str.strip().str.lower()

    def find_col(*labels):
        for label in labels:
            matches = header_vals[header_vals == label].index
            if len(matches):
                return matches[0]
        raise ValueError(f"Could not find any of {labels} in Lazada Marketplace file.")

    col_product_name = find_col("nama produk", "product name")
    col_sku = find_col("seller sku")
    col_stock = find_col("jumlah stok")
    col_product_id = find_col("product id", "id produk")

    data = raw.iloc[header_row + 1:].copy()
    pid = pd.to_numeric(data[col_product_id], errors="coerce")
    data = data[pid.notna()]

    out = pd.DataFrame({
        "Product Name": data[col_product_name],
        "SKU": data[col_sku].astype(str).str.strip(),
        "Marketplace_Stock": pd.to_numeric(data[col_stock], errors="coerce").fillna(0),
    })
    out["SKU"] = out["SKU"].replace({"nan": np.nan, "None": np.nan, "": np.nan})
    out = out.dropna(subset=["SKU"]).reset_index(drop=True)
    out["SKU_Source"] = "Seller SKU (M)"

    out = out.groupby("SKU", as_index=False).agg({
        "Product Name": "first",
        "SKU_Source": "first",
        "Marketplace_Stock": "sum",
    })
    return out[["SKU", "Product Name", "SKU_Source", "Marketplace_Stock"]]


def parse_marketplace_file_shopify(file) -> pd.DataFrame:
    """Parse a Shopify product export. SKU = column R ('Variant SKU').
    Stock = column U ('Variant Inventory Qty'). Single header row (row 0) —
    no extra template rows — so data rows = every row after the header where
    Variant SKU is non-blank. 'Title' is often blank on variant rows, so
    Product Name may come through empty — that's expected."""
    raw = pd.read_excel(file, header=None, dtype=object)

    header_row = None
    for i in range(min(5, len(raw))):
        row_vals = raw.iloc[i].astype(str).str.strip().str.lower()
        if (row_vals == "variant sku").any() and (row_vals == "variant inventory qty").any():
            header_row = i
            break
    if header_row is None:
        raise ValueError(
            "Could not find the header row in the Shopify Marketplace file "
            "(expected 'Variant SKU' / 'Variant Inventory Qty' labels)."
        )

    header_vals = raw.iloc[header_row].astype(str).str.strip().str.lower()

    def find_col(*labels, required=True):
        for label in labels:
            matches = header_vals[header_vals == label].index
            if len(matches):
                return matches[0]
        if required:
            raise ValueError(f"Could not find any of {labels} in Shopify Marketplace file.")
        return None

    col_sku = find_col("variant sku")
    col_stock = find_col("variant inventory qty")
    col_name = find_col("title", required=False)

    data = raw.iloc[header_row + 1:].copy()
    sku_series = data[col_sku].astype(str).str.strip()
    data = data[(sku_series != "") & (sku_series.str.lower() != "nan")]

    out = pd.DataFrame({
        "Product Name": data[col_name] if col_name is not None else "",
        "SKU": data[col_sku].astype(str).str.strip(),
        "Marketplace_Stock": pd.to_numeric(data[col_stock], errors="coerce").fillna(0),
    })
    out["SKU"] = out["SKU"].replace({"nan": np.nan, "None": np.nan, "": np.nan})
    out = out.dropna(subset=["SKU"]).reset_index(drop=True)
    out["Product Name"] = out["Product Name"].fillna("")
    out["SKU_Source"] = "Variant SKU (R)"

    out = out.groupby("SKU", as_index=False).agg({
        "Product Name": "first",
        "SKU_Source": "first",
        "Marketplace_Stock": "sum",
    })
    return out[["SKU", "Product Name", "SKU_Source", "Marketplace_Stock"]]


def parse_warehouse_file(file) -> pd.DataFrame:
    """Parse the Warehouse export. Expects Internal Reference (SKU) and
    Available quantity columns (matched case-insensitively)."""
    df = pd.read_excel(file, dtype=object)
    cols = {c.strip().lower(): c for c in df.columns}

    def find_col(*candidates):
        for cand in candidates:
            if cand in cols:
                return cols[cand]
        raise ValueError(f"Could not find any of {candidates} in Warehouse file columns: {list(df.columns)}")

    sku_col = find_col("internal reference", "sku")
    qty_col = find_col("available quantity", "quantity", "stock")

    out = pd.DataFrame({
        "SKU": df[sku_col].astype(str).str.strip(),
        "Warehouse_Stock": pd.to_numeric(df[qty_col], errors="coerce").fillna(0),
    })
    out = out.dropna(subset=["SKU"])
    return out.groupby("SKU", as_index=False)["Warehouse_Stock"].sum()


def parse_platform_file(file):
    """Parse the Platform export (csv or xlsx). Platform stock = SiAWMS-1
    quantity ONLY (column K in the standard export) — other location columns
    like MyStock-Location quantity are ignored per business rule."""
    name = getattr(file, "name", "")
    if str(name).lower().endswith(".csv"):
        df = pd.read_csv(file, low_memory=False, dtype=object)
    else:
        df = pd.read_excel(file, dtype=object)

    sku_col = None
    for c in df.columns:
        if c.strip().lower() == "sellersku":
            sku_col = c
            break
    if sku_col is None:
        raise ValueError(f"Could not find 'sellerSKU' column in Platform file columns: {list(df.columns)}")

    qty_col = None
    for c in df.columns:
        if "siawms" in c.strip().lower() and "quantity" in c.strip().lower():
            qty_col = c
            break
    if qty_col is None:
        cols = list(df.columns)
        if len(cols) > 10:
            qty_col = cols[10]  # fallback: column K by position
        else:
            raise ValueError("Could not find 'SiAWMS-1 quantity' column in Platform file.")

    df = df[df[sku_col].notna()].copy()
    df[qty_col] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0)

    out = pd.DataFrame({
        "SKU": df[sku_col].astype(str).str.strip(),
        "Platform_Stock": df[qty_col],
    })
    return out.groupby("SKU", as_index=False)["Platform_Stock"].sum(), qty_col


def parse_exclusion_file(file) -> set:
    """Parses an exclusion list file (SKUs whose stock should never be
    flagged for change). Expects a single column of SKUs, optionally preceded
    by blank rows and/or a header row (e.g. 'SKU', 'Seller SKU', 'Variant SKU')."""
    name = getattr(file, "name", "")
    if str(name).lower().endswith(".csv"):
        raw = pd.read_csv(file, header=None, dtype=object)
    else:
        raw = pd.read_excel(file, header=None, dtype=object)

    if raw.empty:
        return set()

    # .astype(str) leaves real NaNs as NaN (not the string "nan") on pandas
    # 3.x, so use dropna() rather than checking for the literal string "nan".
    col0 = raw.iloc[:, 0].astype(str).str.strip().dropna()

    header_labels = {"sku", "seller sku", "variant sku", "internal reference", "item code", "product sku"}
    return {v for v in col0 if v and v.lower() not in header_labels}


def parse_exclusion_text(text: str) -> set:
    """Parses a free-text box of SKUs (newline, comma, semicolon, or
    whitespace separated) into a set of stripped SKU strings."""
    if not text:
        return set()
    tokens = re.split(r"[,\n\r;\t ]+", text.strip())
    return {t.strip() for t in tokens if t.strip()}


def build_comparison(marketplace_df, warehouse_df=None, platform_df=None, excluded_skus=None) -> pd.DataFrame:
    """
    warehouse_df and platform_df are both optional, but at least one must be
    given:
      - All three given: unchanged 3-way logic — Marketplace and Platform are
        each compared against Warehouse (anchor).
      - Marketplace + Warehouse only: Marketplace is compared against
        Warehouse (anchor). No Platform columns are produced.
      - Marketplace + Platform only: no Warehouse anchor available, so
        Platform is compared directly against Marketplace instead.
    """
    have_wh = warehouse_df is not None
    have_pl = platform_df is not None
    if not have_wh and not have_pl:
        raise ValueError("At least one of Warehouse file or Platform file must be provided.")

    excluded_skus = excluded_skus or set()

    df = marketplace_df.copy()
    if have_wh:
        df = df.merge(warehouse_df, on="SKU", how="left")
        df["Found_in_Warehouse"] = df["Warehouse_Stock"].notna()
        df["Warehouse_Stock"] = df["Warehouse_Stock"].fillna(0)
    if have_pl:
        df = df.merge(platform_df, on="SKU", how="left")
        df["Found_in_Platform"] = df["Platform_Stock"].notna()
        df["Platform_Stock"] = df["Platform_Stock"].fillna(0)

    def status(row, found_col, diff_col):
        if not row[found_col]:
            return "SKU Not Found"
        return "Match" if row[diff_col] == 0 else "Mismatch"

    status_cols = []
    diff_cols = []
    output_cols = ["SKU", "Product Name", "SKU_Source", "Marketplace_Stock"]

    if have_pl:
        output_cols.append("Platform_Stock")
    if have_wh:
        output_cols.append("Warehouse_Stock")
        output_cols.append("Correct_Stock (Warehouse anchor)")
        df["Correct_Stock (Warehouse anchor)"] = df["Warehouse_Stock"]

    if have_wh:
        df["MP_vs_WH_Diff"] = (df["Marketplace_Stock"] - df["Warehouse_Stock"]).astype(int)
        df["Marketplace_Status"] = df.apply(lambda r: status(r, "Found_in_Warehouse", "MP_vs_WH_Diff"), axis=1)
        status_cols.append("Marketplace_Status")
        diff_cols.append("MP_vs_WH_Diff")

    if have_wh and have_pl:
        df["PLT_vs_WH_Diff"] = (df["Platform_Stock"] - df["Warehouse_Stock"]).astype(int)
        df["Platform_Status"] = df.apply(lambda r: status(r, "Found_in_Platform", "PLT_vs_WH_Diff"), axis=1)
        status_cols.append("Platform_Status")
        diff_cols.append("PLT_vs_WH_Diff")
    elif have_pl and not have_wh:
        df["MP_vs_PLT_Diff"] = (df["Marketplace_Stock"] - df["Platform_Stock"]).astype(int)
        df["Platform_Status"] = df.apply(lambda r: status(r, "Found_in_Platform", "MP_vs_PLT_Diff"), axis=1)
        status_cols.append("Platform_Status")
        diff_cols.append("MP_vs_PLT_Diff")

    df["Discrepancy_Count"] = sum((df[c] != "Match").astype(int) for c in status_cols)
    df["Excluded"] = df["SKU"].isin(excluded_skus)
    df["Overall_Status"] = np.where(
        df["Excluded"], "Excluded",
        np.where(df["Discrepancy_Count"] == 0, "OK", "Discrepancy"),
    )

    output_cols += diff_cols + status_cols + ["Discrepancy_Count", "Excluded", "Overall_Status"]
    df = df[output_cols]
    return df.sort_values(["Discrepancy_Count", "SKU"], ascending=[False, True]).reset_index(drop=True)


def build_result_file(comparison_df) -> pd.DataFrame:
    """Only SKUs with a discrepancy that are NOT excluded — excluded SKUs are
    never flagged for a stock change."""
    return comparison_df[comparison_df["Overall_Status"] == "Discrepancy"].reset_index(drop=True)


def to_formatted_excel(df: pd.DataFrame, sheet_name: str = "Sheet1") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
    cell_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    match_fill = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
    mismatch_fill = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")
    notfound_fill = PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C")
    excluded_fill = PatternFill("solid", start_color="D9D9D9", end_color="D9D9D9")

    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    status_cols = [c for c in df.columns if c in ("Marketplace_Status", "Platform_Status", "Overall_Status")]
    status_col_idx = {c: df.columns.get_loc(c) + 1 for c in status_cols}

    for _, row in df.iterrows():
        ws.append(list(row))
        r = ws.max_row
        for c_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=r, column=c_idx)
            cell.font = cell_font
            cell.border = border
        for col_name, c_idx in status_col_idx.items():
            val = str(row[col_name])
            cell = ws.cell(row=r, column=c_idx)
            if "Excluded" in val:
                cell.fill = excluded_fill
            elif "Not Found" in val:
                cell.fill = notfound_fill
            elif "Mismatch" in val or "Discrepancy" in val:
                cell.fill = mismatch_fill
            elif "Match" in val or "OK" in val:
                cell.fill = match_fill

    for i, col in enumerate(df.columns, start=1):
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str)]) if len(df) else len(str(col))
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 10), 45)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def style_status(val):
    if val == "Excluded":
        return "background-color:#D9D9D9"
    if val in ("Mismatch", "Discrepancy"):
        return "background-color:#FFC7CE"
    if val == "SKU Not Found":
        return "background-color:#FFEB9C"
    if val in ("Match", "OK"):
        return "background-color:#C6EFCE"
    return ""


def apply_status_styles(df, subset):
    """pandas >=2.1 renamed Styler.applymap -> Styler.map (applymap removed in
    pandas 3.x). Use .map when available, fall back to .applymap on old pandas."""
    styler = df.style
    if hasattr(styler, "map"):
        return styler.map(style_status, subset=subset)
    return styler.applymap(style_status, subset=subset)


# ============================================================================
# UI
# ============================================================================
st.title("📦 Stock Validation Automation")
st.caption(
    "Cross-checks stock between **Marketplace**, **Platform**, and **Warehouse** files. "
    "Marketplace defines the SKU universe. Warehouse is treated as the source of truth."
)

with st.expander("ℹ️ How this works", expanded=False):
    st.markdown(
        """
        1. **SKU lookup** always comes from the Marketplace file.
           - **Shopee:** column **F (SKU)**; if that's blank, column **E (Parent SKU)** is used instead.
           - **TikTok Shop:** column **H (SKU Penjual)**.
           - **Lazada:** column **M (Seller SKU)**.
           - **Shopify:** column **R (Variant SKU)**.
        2. For each Marketplace SKU, we look up the matching stock in the Warehouse and
           Platform files.
           - **Marketplace stock:** Shopee → `Stock` column; TikTok → column **G (Kuantitas)**;
             Lazada → column **H (Jumlah Stok)**; Shopify → column **U (Variant Inventory Qty)**.
        3. **Warehouse is the anchor** whenever it's provided. If Marketplace or Platform
           disagree with Warehouse, Warehouse's number is treated as correct. You can
           upload just Platform, just Warehouse, or both — with only one, Marketplace is
           compared directly against whichever one you provide.
        4. Any SKU with at least one mismatch is compiled into the **Discrepancy Result**
           file. The **Working Process** file contains the full comparison for every SKU
           (useful as an audit trail).
        5. **Exclusion list (optional):** SKUs pasted into the text box or uploaded as a
           file are marked **Excluded** and are never included in the Discrepancy Result,
           even if their numbers don't match — use this for SKUs whose stock should stay
           untouched.
        """
    )

st.divider()

marketplace_platform = st.selectbox(
    "🛍️ Marketplace platform", options=["Shopee", "TikTok Shop", "Lazada", "Shopify"], key="mp_platform"
)
mp_label_map = {
    "Shopee": "🛒 Marketplace file (Shopee export)",
    "TikTok Shop": "🛒 Marketplace file (TikTok Shop export)",
    "Lazada": "🛒 Marketplace file (Lazada export)",
    "Shopify": "🛒 Marketplace file (Shopify export)",
}
mp_label = mp_label_map[marketplace_platform]

col1, col2, col3 = st.columns(3)
with col1:
    marketplace_file = st.file_uploader(mp_label, type=["xlsx", "xls"], key="mp")
with col2:
    platform_file = st.file_uploader(
        "🔗 Platform file (optional)", type=["csv", "xlsx", "xls"], key="plt"
    )
with col3:
    warehouse_file = st.file_uploader(
        "🏭 Warehouse file (optional, anchor)", type=["xlsx", "xls"], key="wh"
    )
st.caption(
    "Upload Platform, Warehouse, or both. With both: Warehouse is the anchor and Platform is "
    "checked against it too. With only one: Marketplace is compared directly against whichever one you provide."
)

with st.expander("🚫 Exclusion list (optional) — SKUs whose stock will never be flagged for change", expanded=False):
    excl_col1, excl_col2 = st.columns(2)
    with excl_col1:
        exclusion_text = st.text_area(
            "Paste SKUs to exclude",
            placeholder="One per line, or comma-separated\ne.g.\nProductsample1\nProductsample7, Productsample12",
            key="excl_text",
            height=120,
        )
    with excl_col2:
        exclusion_file = st.file_uploader(
            "Or upload an exclusion list file",
            type=["csv", "xlsx", "xls"],
            key="excl_file",
            help="Single column of SKUs, optionally with a header row labelled 'SKU'.",
        )

run = st.button("▶️ Run Stock Validation", type="primary", use_container_width=False)

if run:
    if not marketplace_file:
        st.error("Please upload a Marketplace file before running validation.")
        st.stop()
    if not (platform_file or warehouse_file):
        st.error("Please upload at least one of Platform file or Warehouse file before running validation.")
        st.stop()

    try:
        with st.spinner("Parsing files..."):
            if marketplace_platform == "TikTok Shop":
                mp_df = parse_marketplace_file_tiktok(marketplace_file)
            elif marketplace_platform == "Lazada":
                mp_df = parse_marketplace_file_lazada(marketplace_file)
            elif marketplace_platform == "Shopify":
                mp_df = parse_marketplace_file_shopify(marketplace_file)
            else:
                mp_df = parse_marketplace_file_shopee(marketplace_file)
            wh_df = parse_warehouse_file(warehouse_file) if warehouse_file is not None else None
            if platform_file is not None:
                plt_df, qty_cols_used = parse_platform_file(platform_file)
            else:
                plt_df, qty_cols_used = None, None

            excluded_skus = parse_exclusion_text(exclusion_text)
            if exclusion_file is not None:
                excluded_skus |= parse_exclusion_file(exclusion_file)

        with st.spinner("Comparing stock..."):
            comparison = build_comparison(mp_df, wh_df, plt_df, excluded_skus=excluded_skus)
            result = build_result_file(comparison)

        st.session_state["comparison"] = comparison
        st.session_state["result"] = result
        st.session_state["qty_cols_used"] = qty_cols_used
        st.session_state["excluded_count"] = len(excluded_skus)
        st.session_state["has_warehouse"] = warehouse_file is not None
        st.session_state["has_platform"] = platform_file is not None
        st.success("Validation complete.")
    except Exception as e:
        st.error(f"Something went wrong while processing the files: {e}")
        st.stop()

if "comparison" in st.session_state:
    comparison = st.session_state["comparison"]
    result = st.session_state["result"]
    qty_cols_used = st.session_state.get("qty_cols_used", [])
    excluded_count = st.session_state.get("excluded_count", 0)
    has_warehouse = st.session_state.get("has_warehouse", True)
    has_platform = st.session_state.get("has_platform", True)

    st.divider()
    total_sku = len(comparison)
    ok_sku = (comparison["Overall_Status"] == "OK").sum()
    disc_sku = (comparison["Overall_Status"] == "Discrepancy").sum()
    excl_sku = (comparison["Overall_Status"] == "Excluded").sum()

    metric_cols = st.columns(4 + int(has_warehouse) + int(has_platform))
    metric_cols[0].metric("Total SKU", total_sku)
    metric_cols[1].metric("✅ Matched", int(ok_sku))
    metric_cols[2].metric("⚠️ Discrepancies", int(disc_sku), delta=f"{disc_sku/total_sku:.0%} of SKUs" if total_sku else None, delta_color="inverse")
    metric_cols[3].metric("🚫 Excluded", int(excl_sku))
    next_col = 4
    if has_warehouse:
        mp_mismatch = (comparison["Marketplace_Status"] != "Match").sum()
        metric_cols[next_col].metric("Marketplace mismatches", int(mp_mismatch))
        next_col += 1
    if has_platform:
        plt_mismatch = (comparison["Platform_Status"] != "Match").sum()
        metric_cols[next_col].metric("Platform mismatches", int(plt_mismatch))

    if has_warehouse and has_platform:
        st.caption(f"Platform stock computed from column: {qty_cols_used}")
    elif has_platform:
        st.caption(f"Platform stock computed from column: {qty_cols_used}. No Warehouse file provided — Platform is compared directly against Marketplace.")
    else:
        st.caption("No Platform file provided — Marketplace is compared directly against Warehouse (anchor).")
    if excluded_count:
        st.caption(f"{excluded_count} SKU(s) in the exclusion list — their stock will never be flagged for change, even if numbers differ.")

    tab1, tab2 = st.tabs(["⚠️ Discrepancy Result", "📋 Working Process (all SKU)"])

    with tab1:
        st.write(f"**{len(result)} SKU** with at least one discrepancy.")
        if len(result):
            status_subset = [c for c in ["Marketplace_Status", "Platform_Status", "Overall_Status"] if c in result.columns]
            st.dataframe(
                apply_status_styles(result, status_subset),
                use_container_width=True,
                height=450,
            )
            st.download_button(
                "⬇️ Download Discrepancy Result (.xlsx)",
                data=to_formatted_excel(result, "Discrepancies"),
                file_name="Stock_Validation_Discrepancy_Result.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=False,
            )
        else:
            st.success("No discrepancies found — all SKUs match across sources. 🎉")

    with tab2:
        st.write(f"Full comparison for all **{total_sku} SKU** (audit trail).")
        status_subset_all = [c for c in ["Marketplace_Status", "Platform_Status", "Overall_Status"] if c in comparison.columns]
        st.dataframe(
            apply_status_styles(comparison, status_subset_all),
            use_container_width=True,
            height=500,
        )
        st.download_button(
            "⬇️ Download Working Process file (.xlsx)",
            data=to_formatted_excel(comparison, "Working Process"),
            file_name="Stock_Validation_Working_Process.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=False,
        )
else:
    st.info("Upload the three files above and click **Run Stock Validation** to get started.")
